import openpyxl
from openpyxl import Workbook
import sqlite3
import pandas as pd
import sqlite3
import os
from datetime import datetime, date
from math import sin, cos, sqrt, atan2, radians
import csv
from datetime import timedelta
path = '/Users/gimduyeon/Google Drive File Stream/My Drive/gitlab/drm-project/data/'
rawdatapath = '/Users/gimduyeon/Dropbox (KAIST Dr.M)/KAIST Dr.M의 팀 폴더/공유/rawdata/'

files = sorted(os.listdir(rawdatapath))
testStu = [int(i) for i in files if i != '.DS_Store']

f = open(path + '/classroomap_drm.csv', 'r')
# AP data preprocessing
aplist = [r for r in csv.reader(f)]
aplist = [[i[0], i[1][1:-1].replace('"', '').split(",")] for i in aplist]

# get the classes' schedule data from scheduledatebase db file
conn = sqlite3.connect(path + 'SCHEDULEDATABASE.db')
cur = conn.cursor()
cur.execute('SELECT * FROM SCHEDULETABLE')
classschedule = [list(row) for row in cur.fetchall()]

cur.execute('SELECT * FROM SCHEDULETABLE')
classschedule_tmp = [list(row) for row in cur.fetchall()]
conn.close()

data_log_format = '%Y.%m.%d_%H.%M.%S'
num_to_day = {1: 'Mo', 2: 'Tu', 3: 'We', 4: 'Th', 5: 'Fr', 6: 'Sa', 7: 'Su'}
num_to_act = {2: 'VEHICLE', 3: 'BYCYCLE', 4: 'FOOT', 5: 'RUNNING', 6: 'STILL', 7: 'WALKING', 8: 'UNKNOWN'}


for i in classschedule_tmp:
    [i.append(j[1]) for j in aplist if j[0] == i[3]]
classschedule_pd = pd.DataFrame(classschedule_tmp, columns=['class', 'major', 'part', 'loc', 't1', 't2', 'BSSID'])


for t in testStu[59:]:
    print(t)
    excel_document = openpyxl.load_workbook(path+'attendanceResultswithGPS/'+str(t)+'.xlsx')
    sheets = excel_document.get_sheet_names() #get sheet name. sheet names are class names

    attendData = {}
    for sh in sheets:  # 한 시트당 수업 하나
        print(sh)
        sheet = excel_document.get_sheet_by_name(sh)
        sheet.cell(row=1, column=3).value = '입실시간'
        sheet.cell(row=1, column=4).value = '퇴실시간'
        allrows = sheet.rows
        attendData[sh] = allrows
        for r in attendData[sh]:
            entranceTime = 0
            leaveTime = 0
            if r[1].value != sh and r[1].value == 'gps출석':  # gps출석인 경우만 확인
                print(r[0].value)
                actiLogs = []
                actiLogsInClass = []

                attendDate = datetime.strptime(r[0].value, '%Y.%m.%d')
                filerange = [attendDate - timedelta(days=1), attendDate, attendDate + timedelta(days=1)]
                files = sorted([i for i in os.listdir(rawdatapath + str(t)) if i != '.DS_Store' and datetime.strptime(i.split("_")[1], '%Y.%m.%d') in filerange])

                # 출석한 날짜 activity log 추출
                for i in files:
                    conn = sqlite3.connect(rawdatapath + str(t) + '/' + i)
                    cur = conn.cursor()
                    try:
                        cur.execute('SELECT * FROM ACTIVITYTABLE')
                    except sqlite3.OperationalError as e:
                        print(e)
                    else:
                        [actiLogs.append(row) for row in cur.fetchall()]
                    conn.close()

                for i in actiLogs:

                    logtime = datetime.strptime(i[1], data_log_format)
                    logday = num_to_day[logtime.isoweekday()]

                    classData = [i for i in classschedule if i[0] == sheet['B1'].value][0]  # 해당 수업 메타 데이터 가져오기
                    classday = {i.split(".")[0]: [i.split("~")[0].split(".")[1], i.split("~")[1].split(".")[1]] for i in
                                classData[4:] if i != ''}

                    if logday in classday.keys():
                        classStart = datetime.strptime(classday[logday][0], '%H:%M').time()
                        classEnd = datetime.strptime(classday[logday][1], '%H:%M').time()

                        if classStart.minute < 15:
                            classStartORGN = classStart
                            classStart = classStart.replace(hour=classStart.hour - 1, minute=classStart.minute - 15 + 60)
                        else:
                            classStartORGN = classStart
                            classStart = classStart.replace(minute=classStart.minute - 15)

                        if classEnd.minute >= 45:
                            classEndORGN = classEnd
                            classEnd = classEnd.replace(hour=classEnd.hour + 1, minute=classEnd.minute + 15 - 60)
                        else:
                            classEndORGN = classEnd
                            classEnd = classEnd.replace(minute=classEnd.minute + 15)

                        if classStart < logtime.time() < classEnd:
                            actiLogsInClass.append(i)
                            # print(i[1],end="=")
                            # print(num_to_act[i.index(max(i[2:]))])
                # 중복제거
                count = 0
                actiLogsInClass_set = []
                for i in actiLogsInClass:
                    if count == 0:
                        ts = i[1]
                        actiLogsInClass_set.append(i)
                    else:
                        if ts != i[1]:
                            actiLogsInClass_set.append(i)
                            ts = i[1]
                    count += 1

                count = 0
                for i in actiLogsInClass_set:
                    if num_to_act[i.index(max(i[2:]))] == 'STILL':
                        count += 1
                    else:
                        count = 0

                    if count == 3:
                        k = actiLogsInClass_set.index(i) - 2
                        entranceTime = datetime.strptime(actiLogsInClass_set[k][1].split("_")[1], '%H.%M.%S').time()

                        break
                if type(entranceTime) != int and entranceTime < classStartORGN:
                    entranceTime = classStartORGN
                elif type(entranceTime) == int:
                    entranceTime = 'activity error'

                actiLogsInClass_set.reverse()
                count = 0
                for i in actiLogsInClass_set:
                    if num_to_act[i.index(max(i[2:]))] == 'STILL':
                        count += 1
                    else:
                        count = 0

                    if count == 3:
                        k = actiLogsInClass_set.index(i) - 2
                        leaveTime = datetime.strptime(actiLogsInClass_set[k][1].split("_")[1], '%H.%M.%S').time()
                        break
                if type(leaveTime) != int and leaveTime > classEndORGN or \
                                [num_to_act[i.index(max(i[2:]))] for i in actiLogsInClass_set[:3]] == ['STILL', 'STILL',
                                                                                                       'STILL']:
                    leaveTime = classEndORGN
                elif type(leaveTime) == int:
                    leaveTime = 'activity error'
                # print(classStartORGN,end=' - ')
                # print(classEndORGN)
                # print(entranceTime,end=' ~ ')
                # print(leaveTime)

                sheet.cell(row=r[0].row, column=3).value = entranceTime
                sheet.cell(row=r[0].row, column=4).value = leaveTime

    print("#" * 10)
    attendData2= {}
    sheets = excel_document.get_sheet_names() #get sheet name. sheet names are class names

    for sh in sheets:  # 한 시트당 수업 하나
        print(sh)
        sheet = excel_document.get_sheet_by_name(sh)
        allrows = sheet.rows
        attendData2[sh] = allrows
        for r in attendData2[sh]:
            entranceTime = 0
            leaveTime = 0

            if r[1].value != sh and r[1].value == '출석':  # wifi출석인 경우만 확인
                print(r[0].value)
                actiLogs = []
                actiLogsInClass = []
                wifiLogs = []
                wifiLogsInClass = []

                attendDate = datetime.strptime(r[0].value, '%Y.%m.%d')
                filerange = [attendDate - timedelta(days=1), attendDate, attendDate + timedelta(days=1)]
                files = sorted([i for i in os.listdir(rawdatapath + str(t))
                                if i != '.DS_Store' and datetime.strptime(i.split("_")[1], '%Y.%m.%d') in filerange])

                # 출석한 날짜 activity log 추출
                for i in files:
                    conn = sqlite3.connect(rawdatapath + str(t) + '/' + i)
                    cur = conn.cursor()
                    try:
                        cur.execute('SELECT * FROM ACTIVITYTABLE')
                    except sqlite3.OperationalError as e:
                        print(e)
                    else:
                        [actiLogs.append(row) for row in cur.fetchall()]
                    try:
                        cur.execute('SELECT * FROM HARDWARETABLE')
                    except sqlite3.OperationalError as e:
                        print(e)
                    else:
                        [wifiLogs.append(row) for row in cur.fetchall() if row[2] == 'WIFI1' or row[2] == 'WIFI2']

                    conn.close()

                for i in actiLogs:

                    logtime = datetime.strptime(i[1], data_log_format)
                    logday = num_to_day[logtime.isoweekday()]

                    classData = [i for i in classschedule if i[0] == sheet['B1'].value][0]  # 해당 수업 메타 데이터 가져오기
                    classday = {i.split(".")[0]: [i.split("~")[0].split(".")[1], i.split("~")[1].split(".")[1]] for i in
                                classData[4:] if i != ''}

                    if logday in classday.keys():
                        classStart = datetime.strptime(classday[logday][0], '%H:%M').time()
                        classEnd = datetime.strptime(classday[logday][1], '%H:%M').time()

                        if classStart.minute < 15:
                            classStartORGN = classStart
                            classStart = classStart.replace(hour=classStart.hour - 1, minute=classStart.minute - 15 + 60)
                        else:
                            classStartORGN = classStart
                            classStart = classStart.replace(minute=classStart.minute - 15)

                        if classEnd.minute >= 45:
                            classEndORGN = classEnd
                            classEnd = classEnd.replace(hour=classEnd.hour + 1, minute=classEnd.minute + 15 - 60)
                        else:
                            classEndORGN = classEnd
                            classEnd = classEnd.replace(minute=classEnd.minute + 15)

                        if classStart < logtime.time() < classEnd:
                            actiLogsInClass.append(i)
                            # print(i[1],end="=")
                            # print(num_to_act[i.index(max(i[2:]))])

                for i in wifiLogs:
                    logtime = datetime.strptime(i[1], data_log_format)
                    logday = num_to_day[logtime.isoweekday()]
                    if logday in classday.keys():
                        if classStart < logtime.time() < classEnd:
                            wifiLogsInClass.append(i)

                # activity 중복제거
                count = 0
                actiLogsInClass_set = []
                for i in actiLogsInClass:

                    if count == 0:
                        ts = i[1]
                        actiLogsInClass_set.append(i)
                    else:
                        if ts != i[1]:
                            actiLogsInClass_set.append(i)
                            ts = i[1]
                    count += 1

                # wifi로 출석 체크한 경우에는 wifi가 도는 시간을 확인 해서 entrance 시간 확인
                # 해당 wifi의 ap가 해당 강의실 내부에 있는 것인지 확인해야함
                # 해당 강의실의 ap 리스트 중 가장 먼저 잡히는 것으로 도착시간 확인
                classroom_ap = classschedule_pd[classschedule_pd['class'] == sheet['B1'].value].BSSID.values[0]
                for i in wifiLogsInClass:

                    if i[4] != None and i[4][:-2] in [i[:-2] for i in classroom_ap]:  # tanimoto
                        entranceTime = datetime.strptime(i[1].split("_")[1], '%H.%M.%S').time()
                        break
                if type(entranceTime) != int and entranceTime < classStartORGN:
                    entranceTime = classStartORGN
                elif type(entranceTime) == int:
                    entranceTime = 'wifi 스캔에러'
                # 나간 시간 체크
                actiLogsInClass_set.reverse()
                count = 0
                for i in actiLogsInClass_set:
                    if num_to_act[i.index(max(i[2:]))] == 'STILL':
                        count += 1
                    else:
                        count = 0

                    if count == 3:
                        k = actiLogsInClass_set.index(i) - 2
                        leaveTime = datetime.strptime(actiLogsInClass_set[k][1].split("_")[1], '%H.%M.%S').time()
                        break
                if type(leaveTime) != int and leaveTime > classEndORGN or \
                                [num_to_act[i.index(max(i[2:]))] for i in actiLogsInClass_set[:3]] == ['STILL', 'STILL',
                                                                                                       'STILL']:
                    leaveTime = classEndORGN
                elif type(leaveTime) == int:
                    leaveTime = 'activity error'
                # print(classStartORGN,end=' - ')
                # print(classEndORGN)
                # print(entranceTime,end=' ~ ')
                # print(leaveTime)
                sheet.cell(row=r[0].row, column=3).value = entranceTime
                sheet.cell(row=r[0].row, column=4).value = leaveTime


    excel_document.save(path+'entranceAndLeaveResults/'+str(t) + '_el.xlsx')
    excel_document.close()